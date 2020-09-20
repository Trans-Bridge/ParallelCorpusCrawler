import pickle
import logging

from openpyxl import load_workbook, Workbook
from typing import List, Tuple
from tqdm import tqdm


class BaseCrawler(object):

    def __init__(self,
                 src_kw: List[str],
                 tgt_kw: List[str]) -> None:

        self.src_kw = src_kw
        self.tgt_kw = tgt_kw

        self.error_list = {
            "src": [],
            "tgt": []
        }
        self.corpus = []

    @classmethod
    def from_excel(cls,
                   excel_path: str,
                   src_column: int = 0,
                   tgt_column: int = 1,
                   skip_first_column: bool = True,
                   **kwargs):
        src_words, tgt_words = [], []
        sheet = load_workbook(filename=excel_path, read_only=True).active

        for i, row in tqdm(enumerate(sheet.iter_rows())):
            if not i and skip_first_column:
                continue
            src = row[src_column].value
            tgt = row[tgt_column].value

            if src and tgt:
                src_words.append(str(src))
                tgt_words.append(str(tgt))

        return cls(src_words, tgt_words, **kwargs)

    @classmethod
    def from_checkpoint(cls, checkpoint_path: str):
        obj = pickle.loads(checkpoint_path)
        assert isinstance(obj, cls)

    def crawl(self):
        for i, seeds in tqdm(enumerate(zip(self.src_kw, self.tgt_kw))):
            for seed, tp in zip(seeds, ["src", "tgt"]):
                try:
                    for pair in self._crawl_single(seed, type=tp):
                        self.corpus.append(pair)
                except Exception as e:
                    logging.log(
                        level=logging.WARNING, msg="Crawl with seed %s error. Error message: %s" % (seed, e))
                    self.error_list[tp].append(seed)

    def save(self, file_path: str):
        wb = Workbook()
        sheet = wb.active

        sheet.cell(row=1, column=1, value="Source Sentence")
        sheet.cell(row=1, column=2, value="Target Sentence")

        for i, (src, tgt) in enumerate(self.corpus):
            i += 2
            sheet.cell(row=i, column=1, value=src)
            sheet.cell(row=i, column=2, value=tgt)

        wb.save(file_path)

    def _crawl_single(self, seed: str, type: str = "src") -> Tuple[str, str]:
        raise NotImplementedError
